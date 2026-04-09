"""
csv_to_excel.py
---------------
build_excel(profile, companies, projects) -> openpyxl.Workbook

md_to_excel.py から呼び出される Excel 生成モジュール。
format.csv と同じレイアウトで SkillSheet_Manpyo.xlsx を生成する。

列構成 (A〜S, 1-indexed):
  A(1):No  B(2):開始  C(3):-  D(4):終了  E(5):業務内容
  F(6):役割/規模  G(7):使用言語  H(8):DB  I(9):サーバOS
  J(10):FW・MW  K-Q(11-17):担当工程7列
"""

from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# カラー定数
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
CLR_HEADER_BG   = "1F497D"   # 濃紺（スキルシートタイトル）
CLR_HEADER_FG   = "FFFFFF"   # 白文字
CLR_SECTION_BG  = "4F81BD"   # 中青（セクションヘッダー）
CLR_COL_HDR_BG  = "DBE5F1"   # 薄青（列ヘッダー）
CLR_PROJECT_BG  = "EEF3FA"   # 極薄青（プロジェクト奇数行）
CLR_PHASE_HL    = "FFC000"   # オレンジ（担当工程●）
CLR_BORDER      = "A0A0A0"   # グレー罫線


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# スタイルファクトリ
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def _font(bold=False, color="000000", size=10):
    return Font(bold=bold, color=color, size=size, name="Meiryo UI")


def _align(h="left", v="top", wrap=True):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(style="thin"):
    s = Side(style=style, color=CLR_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


def _thick_border():
    thick = Side(style="medium", color="4F81BD")
    thin  = Side(style="thin",   color=CLR_BORDER)
    return Border(left=thick, right=thick, top=thin, bottom=thin)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# セル操作ヘルパー
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _set(ws, row, col, value="", fill=None, font=None, align=None, border=None):
    """1セルに値・スタイルを設定する。"""
    cell = ws.cell(row=row, column=col, value=value)
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = border
    return cell


def _merge(ws, r1, c1, r2, c2):
    """セル結合（結合後に左上セルを返す）。"""
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
    return ws.cell(row=r1, column=c1)


def _fill_range(ws, r1, c1, r2, c2, fill, font=None, align=None, border=None):
    """矩形範囲の全セルにスタイルを適用する（結合用）。"""
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            _set(ws, r, c, fill=fill, font=font, align=align, border=border)


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# カラム幅設定
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
COL_WIDTHS = {
    1:  5,    # A: No.
    2: 11,    # B: 開始
    3:  2,    # C: (空)
    4:  2,    # D: -
    5: 11,    # E: 終了
    6: 55,    # F: 業務内容（メイン）
    7: 11,    # G: 役割/規模
    8: 14,    # H: 使用言語
    9:  9,    # I: DB
    10: 18,   # J: サーバOS
    11: 20,   # K: FW・MW・ツール等
    12:  7,   # L: 要件定義
    13:  7,   # M: 基本設計
    14:  7,   # N: 詳細設計
    15:  7,   # O: 実装・単体
    16:  7,   # P: 結合テスト
    17:  7,   # Q: 総合テスト
    18:  7,   # R: 保守・運用
}

# 担当工程列（L〜R = 12〜18）
PHASE_COLS   = list(range(12, 19))
PHASE_LABELS = ["要件定義","基本設計","詳細設計","実装・単体","結合テスト","総合テスト","保守・運用"]

NCOLS = 18  # A〜R


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# プロフィールセクション
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _write_profile(ws, profile, companies):
    """
    1〜(9+len(companies)+2) 行目のプロフィール・職務経歴ヘッダーを書き込む。
    最後に書き込んだ行番号を返す。
    """
    hdr_fill   = _fill(CLR_HEADER_BG)
    hdr_font   = _font(bold=True, color=CLR_HEADER_FG, size=11)
    lbl_fill   = _fill(CLR_SECTION_BG)
    lbl_font   = _font(bold=True, color=CLR_HEADER_FG)
    val_fill   = _fill("FFFFFF")
    val_font   = _font()
    val_align  = _align(h="left", v="center", wrap=False)
    bdr        = _border()

    # ── 行1: タイトル ─────────────────────────────────────────
    row = 1
    _fill_range(ws, row, 1, row, NCOLS, fill=hdr_fill)
    cell = _merge(ws, row, 1, row, NCOLS)
    cell.value     = "スキルシート"
    cell.fill      = hdr_fill
    cell.font      = Font(bold=True, color=CLR_HEADER_FG, size=14, name="Meiryo UI")
    cell.alignment = _align(h="center", v="center", wrap=False)

    # ── 行2〜5: プロフィール項目 ──────────────────────────────
    profile_rows = [
        ("技術者名", profile.get("氏名",""),     "所　　属", profile.get("所属","")),
        ("年　　齢", profile.get("年齢",""),      "性　　別", profile.get("性別","")),
        ("最 寄 駅", profile.get("最寄駅",""),    "学　　歴", profile.get("学歴","")),
        ("稼動開始日",profile.get("稼動開始",""), "資　　格", profile.get("資格","")),
    ]
    for i, (lbl1, val1, lbl2, val2) in enumerate(profile_rows):
        row = 2 + i
        # 左ラベル (A〜B結合)
        _fill_range(ws, row, 1, row, 2, fill=lbl_fill)
        c = _merge(ws, row, 1, row, 2)
        c.value = lbl1; c.fill = lbl_fill; c.font = lbl_font
        c.alignment = _align(h="center", v="center", wrap=False)
        # 左値 (C〜H結合)
        _fill_range(ws, row, 3, row, 8, fill=val_fill)
        c = _merge(ws, row, 3, row, 8)
        c.value = val1; c.fill = val_fill; c.font = val_font
        c.alignment = val_align; c.border = bdr
        # 右ラベル (I〜J結合)
        _fill_range(ws, row, 9, row, 10, fill=lbl_fill)
        c = _merge(ws, row, 9, row, 10)
        c.value = lbl2; c.fill = lbl_fill; c.font = lbl_font
        c.alignment = _align(h="center", v="center", wrap=False)
        # 右値 (K〜R結合)
        _fill_range(ws, row, 11, row, NCOLS, fill=val_fill)
        c = _merge(ws, row, 11, row, NCOLS)
        c.value = val2; c.fill = val_fill; c.font = val_font
        c.alignment = val_align; c.border = bdr

    # ── 行6: 得意技術 ────────────────────────────────────────
    row = 6
    _fill_range(ws, row, 1, row, 2, fill=lbl_fill)
    c = _merge(ws, row, 1, row, 2)
    c.value = "得意技術"; c.fill = lbl_fill; c.font = lbl_font
    c.alignment = _align(h="center", v="center", wrap=False)
    _fill_range(ws, row, 3, row, NCOLS, fill=val_fill)
    c = _merge(ws, row, 3, row, NCOLS)
    c.value = profile.get("得意技術",""); c.fill = val_fill; c.font = val_font
    c.alignment = val_align; c.border = bdr

    # ── 行7: 得意業務 ────────────────────────────────────────
    row = 7
    _fill_range(ws, row, 1, row, 2, fill=lbl_fill)
    c = _merge(ws, row, 1, row, 2)
    c.value = "得意業務"; c.fill = lbl_fill; c.font = lbl_font
    c.alignment = _align(h="center", v="center", wrap=False)
    _fill_range(ws, row, 3, row, NCOLS, fill=val_fill)
    c = _merge(ws, row, 3, row, NCOLS)
    c.value = profile.get("得意業務",""); c.fill = val_fill; c.font = val_font
    c.alignment = val_align; c.border = bdr

    # ── 行8: 職務経歴ヘッダー ────────────────────────────────
    row = 8
    sec_fill  = _fill(CLR_SECTION_BG)
    sec_font  = _font(bold=True, color=CLR_HEADER_FG)
    sec_align = _align(h="center", v="center", wrap=False)
    _fill_range(ws, row, 1, row, NCOLS, fill=sec_fill)
    headers_8 = {
        1: ("職務経歴", 1, 2), 3: ("No.", 3, 3), 4: ("企業名", 4, 6),
        7: ("期間", 7, 8), 9: ("契約・雇用形態", 9, 10),
        11: ("担当職種", 11, 12), 13: ("事業内容", 13, NCOLS),
    }
    for start_col, (label, c1, c2) in headers_8.items():
        if c1 != c2:
            _fill_range(ws, row, c1, row, c2, fill=sec_fill)
            c = _merge(ws, row, c1, row, c2)
        else:
            c = ws.cell(row=row, column=c1)
        c.value = label; c.fill = sec_fill; c.font = sec_font; c.alignment = sec_align

    # ── 行9〜: 職務経歴データ ─────────────────────────────────
    comp_fill  = _fill("FFFFFF")
    comp_font  = _font()
    comp_align = _align(h="left", v="center", wrap=True)
    for comp in companies:
        row += 1
        ws.row_dimensions[row].height = 28
        _set(ws, row, 3, comp['no'],       fill=comp_fill, font=comp_font,
             align=_align(h="center", v="center"), border=bdr)
        _fill_range(ws, row, 4, row, 6, fill=comp_fill)
        c = _merge(ws, row, 4, row, 6)
        c.value = comp['name']; c.fill = comp_fill; c.font = Font(bold=True, size=10, name="Meiryo UI")
        c.alignment = comp_align; c.border = bdr
        _fill_range(ws, row, 7, row, 8, fill=comp_fill)
        c = _merge(ws, row, 7, row, 8)
        c.value = comp['period']; c.fill = comp_fill; c.font = comp_font
        c.alignment = _align(h="center", v="center"); c.border = bdr
        _fill_range(ws, row, 9, row, 10, fill=comp_fill)
        c = _merge(ws, row, 9, row, 10)
        c.value = comp['type']; c.fill = comp_fill; c.font = comp_font
        c.alignment = _align(h="center", v="center"); c.border = bdr
        _fill_range(ws, row, 11, row, 12, fill=comp_fill)
        c = _merge(ws, row, 11, row, 12)
        c.value = comp['role']; c.fill = comp_fill; c.font = comp_font
        c.alignment = _align(h="center", v="center"); c.border = bdr
        _fill_range(ws, row, 13, row, NCOLS, fill=comp_fill)
        c = _merge(ws, row, 13, row, NCOLS)
        c.value = comp['business']; c.fill = comp_fill; c.font = comp_font
        c.alignment = comp_align; c.border = bdr

    return row


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# プロジェクトセクションヘッダー
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _write_proj_header(ws, row):
    """担当工程ヘッダー2行を書き込み、最後の行番号を返す。"""
    sec_fill  = _fill(CLR_SECTION_BG)
    sec_font  = _font(bold=True, color=CLR_HEADER_FG)
    c_align   = _align(h="center", v="center", wrap=True)
    bdr       = _border()

    # ── 1行目: 列ラベル ──
    row += 1
    ws.row_dimensions[row].height = 38
    _fill_range(ws, row, 1, row, NCOLS, fill=sec_fill)
    labels_r1 = [
        (1,  1, "期間"),
        (6,  6, "業務内容"),
        (7,  7, "役割\n規模"),
        (8,  8, "使用言語"),
        (9,  9, "DB"),
        (10, 10, "サーバOS"),
        (11, 11, "FW・MW\nツール等"),
        (12, 18, "担当工程"),
    ]
    for c1, c2, label in labels_r1:
        if c1 != c2:
            _fill_range(ws, row, c1, row, c2, fill=sec_fill)
            c = _merge(ws, row, c1, row, c2)
        else:
            c = ws.cell(row=row, column=c1)
        c.value = label; c.fill = sec_fill; c.font = sec_font; c.alignment = c_align

    # ── 2行目: 期間詳細と工程ラベル ──
    row += 1
    ws.row_dimensions[row].height = 28
    col_hdr_fill = _fill(CLR_COL_HDR_BG)
    col_hdr_font = _font(bold=True)
    for col, label in zip(PHASE_COLS, PHASE_LABELS):
        c = _set(ws, row, col, label,
                 fill=col_hdr_fill, font=col_hdr_font,
                 align=_align(h="center", v="center", wrap=True), border=bdr)
    # 期間サブヘッダー (A〜E)
    sub_labels = ["No.", "開始", "", "〜", "終了"]
    for i, label in enumerate(sub_labels):
        _set(ws, row, i+1, label, fill=col_hdr_fill, font=col_hdr_font,
             align=_align(h="center", v="center", wrap=False), border=bdr)
    # その他の列をグレーで塗る
    for col in [6, 7, 8, 9, 10, 11]:
        _set(ws, row, col, fill=col_hdr_fill, border=bdr)

    return row


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# プロジェクトデータ行
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def _write_project(ws, row, proj, idx):
    """1プロジェクト分を書き込み、最後の行番号を返す。"""
    bdr       = _border()
    val_font  = _font()
    c_align   = _align(h="center", v="center", wrap=True)
    l_align   = _align(h="left",   v="top",    wrap=True)

    # 奇数/偶数で背景を変える
    bg_hex  = CLR_PROJECT_BG if idx % 2 == 1 else "FFFFFF"
    bg_fill = _fill(bg_hex)

    # ── タイトル行 ──────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 42

    # No.
    _set(ws, row, 1, idx, fill=bg_fill, font=_font(bold=True),
         align=c_align, border=bdr)
    # 開始
    _set(ws, row, 2, proj['start'], fill=bg_fill, font=val_font,
         align=c_align, border=bdr)
    # 〜 (空)
    _set(ws, row, 3, "", fill=bg_fill, border=bdr)
    # -
    _set(ws, row, 4, "-", fill=bg_fill, font=val_font, align=c_align, border=bdr)
    # 終了
    _set(ws, row, 5, proj['end'], fill=bg_fill, font=val_font,
         align=c_align, border=bdr)
    # 業務内容タイトル
    _set(ws, row, 6, proj['title'], fill=bg_fill,
         font=_font(bold=True), align=l_align, border=bdr)
    # 役割/規模
    _set(ws, row, 7, proj['role'], fill=bg_fill, font=val_font,
         align=c_align, border=bdr)
    # 使用言語
    _set(ws, row, 8, proj['lang'], fill=bg_fill, font=val_font,
         align=c_align, border=bdr)
    # DB
    _set(ws, row, 9, proj['db'], fill=bg_fill, font=val_font,
         align=c_align, border=bdr)
    # サーバOS
    _set(ws, row, 10, proj['os'], fill=bg_fill, font=val_font,
         align=c_align, border=bdr)
    # FW・MW
    _set(ws, row, 11, proj['fw'], fill=bg_fill, font=val_font,
         align=c_align, border=bdr)
    # 担当工程 ●/-
    phase_hl_fill = _fill(CLR_PHASE_HL)
    for i, phase_val in enumerate(proj['phases']):
        col  = PHASE_COLS[i]
        fill = phase_hl_fill if phase_val == "●" else bg_fill
        fnt  = _font(bold=True) if phase_val == "●" else val_font
        _set(ws, row, col, phase_val, fill=fill, font=fnt,
             align=c_align, border=bdr)

    # ── 詳細テキスト行 ────────────────────────────────────────
    row += 1
    detail_lines = proj['detail'].count('\n') + 1
    ws.row_dimensions[row].height = max(60, detail_lines * 14)

    # 業務内容テキスト (F列)
    _set(ws, row, 6, proj['detail'], fill=bg_fill, font=val_font,
         align=l_align, border=bdr)
    # 体制 (G列)
    _set(ws, row, 7, proj['team'], fill=bg_fill, font=val_font,
         align=c_align, border=bdr)
    # その他列を塗りつぶす
    for col in [1, 2, 3, 4, 5, 8, 9, 10, 11] + PHASE_COLS:
        _set(ws, row, col, fill=bg_fill, border=bdr)

    # ── 期間行 ────────────────────────────────────────────────
    row += 1
    ws.row_dimensions[row].height = 16
    dur_fill = _fill("D9E1F2")
    # B列に期間テキスト
    _set(ws, row, 2, proj['duration'], fill=dur_fill,
         font=_font(color="1F497D"), align=_align(h="center", v="center", wrap=False))
    for col in [1, 3, 4, 5, 6, 7, 8, 9, 10, 11] + PHASE_COLS:
        _set(ws, row, col, fill=dur_fill)

    return row


# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
# メインエントリポイント
# ━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
def build_excel(profile, companies, projects):
    """
    profile   : dict (md_to_excel.parse_md の戻り値)
    companies : list[dict]
    projects  : list[dict]
    return    : openpyxl.Workbook
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "スキルシート"

    # ── カラム幅設定 ──────────────────────────────────────────
    for col_idx, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── 印刷設定 ──────────────────────────────────────────────
    ws.page_setup.orientation    = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToPage      = True
    ws.page_setup.fitToWidth     = 1
    ws.page_setup.fitToHeight    = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True

    # ── プロフィール ──────────────────────────────────────────
    last_row = _write_profile(ws, profile, companies)

    # ── プロジェクトヘッダー ──────────────────────────────────
    last_row = _write_proj_header(ws, last_row)

    # ── プロジェクトデータ ────────────────────────────────────
    for idx, proj in enumerate(projects, start=1):
        last_row = _write_project(ws, last_row, proj, idx)

    # ── 先頭行を固定 ─────────────────────────────────────────
    ws.freeze_panes = "A2"

    return wb
