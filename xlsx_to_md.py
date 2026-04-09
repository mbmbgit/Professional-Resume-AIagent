"""
xlsx_to_md.py
-------------
SkillSheet_Manbyo.xlsx を読み込み、Skills.md 形式に変換して出力する。

実行方法:
    python3 xlsx_to_md.py
"""

import openpyxl
import re

INPUT_XLSX = "SkillSheet_Manbyo.xlsx"
OUTPUT_MD  = "Skills_from_xlsx.md"

# 担当工程列（L〜R = 12〜18, 0-indexed: 11〜17）
PHASE_LABELS = ["要件定義","基本設計","詳細設計","実装・単体","結合テスト","総合テスト","保守・運用"]


def cell_val(ws, row, col):
    """セル値を文字列で返す（Noneは空文字）。"""
    v = ws.cell(row=row, column=col).value
    return str(v).strip() if v is not None else ""


def read_all_rows(ws):
    """全行を (row_idx, [col_values...]) のリストで返す。"""
    rows = []
    for r in range(1, ws.max_row + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        rows.append(vals)
    return rows


def find_project_rows(rows):
    """
    プロジェクトデータ行を検出する。
    列A にシーケンス番号（整数または数値文字列）がある行をタイトル行とみなす。
    """
    projects = []
    i = 0
    while i < len(rows):
        row = rows[i]
        a = row[0]  # A列: No.
        # No. が整数（または整数文字列）で、F列に業務内容タイトルがある行
        try:
            no = int(a)
        except (TypeError, ValueError):
            i += 1
            continue

        # F列 (idx=5): タイトル
        title = str(row[5]).strip() if row[5] else ""
        if not title or title in ("業務内容", ""):
            i += 1
            continue

        # 開始・終了 (B, E 列 → idx 1, 4)
        start = str(row[1]).strip() if row[1] else ""
        end   = str(row[4]).strip() if row[4] else ""

        # 役割/規模 (G列 → idx 6)
        role_raw = str(row[6]).strip() if row[6] else ""
        role_lines = role_raw.split("\n")
        role = role_lines[0].strip() if role_lines else ""
        size = role_lines[1].strip() if len(role_lines) > 1 else "1名"

        # 技術情報 (H〜K列 → idx 7〜10)
        lang  = str(row[7]).strip() if row[7] else "-"
        db    = str(row[8]).strip() if row[8] else "-"
        os_   = str(row[9]).strip() if row[9] else "-"
        fw    = str(row[10]).strip() if row[10] else "-"

        # 担当工程 (L〜R → idx 11〜17)
        phases = []
        for idx in range(11, 18):
            v = row[idx] if idx < len(row) else None
            phases.append("●" if str(v).strip() == "●" else "-")

        # 次行: 詳細テキスト (F列) & 体制 (G列)
        detail = ""
        team   = ""
        if i + 1 < len(rows):
            next_row = rows[i + 1]
            detail = str(next_row[5]).strip() if next_row[5] else ""
            team   = str(next_row[6]).strip() if next_row[6] else ""

        # 次々行: 期間テキスト (B列)
        duration = ""
        if i + 2 < len(rows):
            dur_row = rows[i + 2]
            duration = str(dur_row[1]).strip() if dur_row[1] else ""

        # 期間文字列の組み立て
        # "（約Xヶ月）" のような文字列から数字部分を取得
        dur_text = re.sub(r"[（(）)]", "", duration).strip()  # 括弧除去
        if start and end and start != end:
            period = f"{start} 〜 {end}"
            if dur_text:
                period += f"（{dur_text}）"
        elif start:
            period = start
            if dur_text:
                period += f"（{dur_text}）"
        else:
            period = ""

        projects.append({
            "no":      no,
            "title":   title,
            "period":  period,
            "role":    role,
            "size":    size,
            "lang":    lang,
            "db":      db,
            "os":      os_,
            "fw":      fw,
            "phases":  phases,
            "detail":  detail,
            "team":    team,
        })

        i += 3  # タイトル行 + 詳細行 + 期間行
    return projects


def read_profile(ws):
    """プロフィール情報を読み取る（行1〜7）。"""
    profile = {}
    # 行2: 技術者名（B列=idx1）、所属（J列=idx9）
    profile["氏名"]     = cell_val(ws, 2, 2)
    profile["所属"]     = cell_val(ws, 2, 10)
    # 行5: 稼動開始日（B列）、資格（J列）
    profile["稼動開始"] = cell_val(ws, 5, 2)
    profile["資格"]     = cell_val(ws, 5, 10)
    # 行6: 得意技術（C列以降、結合されているため複数列試行）
    for col in range(3, 12):
        v = cell_val(ws, 6, col)
        if v:
            profile["得意技術"] = v
            break
    # 行7: 得意業務
    for col in range(3, 12):
        v = cell_val(ws, 7, col)
        if v:
            profile["得意業務"] = v
            break
    return profile


def detail_to_md(detail_text):
    """
    詳細テキスト（≪担当業務≫ / ≪ビジネス上の効果≫）を
    Markdown の **業務内容・詳細機能:** / **ビジネス上の効果:** 形式に変換する。
    """
    if not detail_text:
        return ""

    lines = detail_text.split("\n")
    out_lines = []
    section = None

    for line in lines:
        line = line.rstrip()
        if "≪担当業務≫" in line:
            section = "duty"
            out_lines.append("**業務内容・詳細機能:**  ")
            continue
        elif "≪ビジネス上の効果≫" in line:
            section = "biz"
            out_lines.append("")
            out_lines.append("**ビジネス上の効果:**  ")
            continue

        # 「　・」で始まる行を Markdown リストに変換
        m = re.match(r"^[\s　]*[・･]\s*(.+)", line)
        if m:
            out_lines.append(f"- {m.group(1).strip()}")
        elif line.strip():
            out_lines.append(line)

    return "\n".join(out_lines)


def build_md(profile, projects):
    lines = []

    # ── ヘッダー ────────────────────────────────────────────────
    lines.append("# スキルシート")
    lines.append("")
    lines.append(f"**氏名**: {profile.get('氏名', '萬俵 秀俊（まんぴょう ひでとし）')}  ")
    lines.append("**更新日**: 2026年4月7日  ")
    lines.append("**職種**: フリーランス Pythonエンジニア / 業務自動化・Webスクレイピング・AI連携")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ── 技術スキル概要（固定） ───────────────────────────────────
    lines.append("## ■ 技術スキル概要")
    lines.append("")
    lines.append("| 分野 | 技術・ツール | 経験年数 | 習熟度 |")
    lines.append("| :--- | :--- | :---: | :---: |")
    skill_rows = [
        ("**言語**", "Python 3", "6年", "★★★★★"),
        ("**言語**", "GAS (Google Apps Script)", "3年", "★★★★☆"),
        ("**言語**", "VB.NET", "1年", "★★★☆☆"),
        ("**言語**", "JavaScript / HTML / CSS", "2年", "★★★☆☆"),
        ("**フレームワーク**", "FastAPI", "1年", "★★★☆☆"),
        ("**フレームワーク**", "Flet (PythonGUIフレームワーク)", "1年", "★★★★☆"),
        ("**スクレイピング**", "Selenium", "4年", "★★★★★"),
        ("**スクレイピング**", "BeautifulSoup", "3年", "★★★★☆"),
        ("**ノーコード**", "Octoparse", "4年", "★★★★★"),
        ("**RPA**", "UiPath", "1年", "★★★☆☆"),
        ("**AI API**", "Gemini API / API", "2年", "★★★★☆"),
        ("**AI API**", "Claude Code / API", "1年", "★★★★☆"),
        ("**クラウド**", "GCP (Google Cloud Platform)", "2年", "★★★☆☆"),
        ("**クラウド**", "GitHub Actions / Codespaces", "2年", "★★★☆☆"),
        ("**DB・データ**", "CSV / Excel / Google Spreadsheet", "6年", "★★★★★"),
        ("**DB・データ**", "SPARQL / WikidataクエリAPI", "1年", "★★★☆☆"),
        ("**OS**", "Windows 10 / 11", "6年", "★★★★★"),
        ("**OS**", "Linux (Ubuntu)", "3年", "★★★★☆"),
        ("**OCR**", "Line Clova OCR / pyMuPDF", "2年", "★★★★☆"),
        ("**通知連携**", "Chatwork / Zapier", "2年", "★★★★☆"),
        ("**AIエディタ**", "Windsurf / VS Code", "2年", "★★★★☆"),
    ]
    for row in skill_rows:
        lines.append(f"| {row[0]} | {row[1]} | {row[2]} | {row[3]} |")
    lines.append("")
    lines.append("---")
    lines.append("")

    # ── プロジェクト経歴 ────────────────────────────────────────
    lines.append("## ■ プロジェクト経歴（スキルシート詳細）")
    lines.append("")
    lines.append("> 特記なき場合、体制は **1名 / PG（プログラマ）** として参画。担当工程の「●」は担当したフェーズを示す。")
    lines.append("")

    for proj in projects:
        lines.append(f"### {proj['no']}. {proj['title'].lstrip('■').strip()}")
        lines.append("")
        lines.append("| 項目 | 内容 |")
        lines.append("| :--- | :--- |")
        lines.append(f"| **期間** | {proj['period']} |")
        lines.append(f"| **役割 / 体制** | {proj['role']} / {proj['size']} |")
        lines.append(f"| **使用言語** | {proj['lang']} |")
        if proj['db'] and proj['db'] != '-':
            lines.append(f"| **DB** | {proj['db']} |")
        if proj['os'] and proj['os'] != '-':
            lines.append(f"| **サーバOS** | {proj['os']} |")
        if proj['fw'] and proj['fw'] != '-':
            # 改行区切りをカンマ区切りに
            fw_inline = proj['fw'].replace("\n", ", ")
            lines.append(f"| **FW・MW・ツール等** | {fw_inline} |")
        lines.append("")

        # 業務内容・効果
        md_detail = detail_to_md(proj['detail'])
        if md_detail:
            lines.append(md_detail)
        else:
            lines.append("**業務内容・詳細機能:**  ")
            lines.append("")
            lines.append("**ビジネス上の効果:**  ")
        lines.append("")

        # 担当工程テーブル
        lines.append("| 要件定義 | 基本設計 | 詳細設計 | 実装・単体 | 結合テスト | 総合テスト | 保守・運用 |")
        lines.append("|:---:|:---:|:---:|:---:|:---:|:---:|:---:|")
        lines.append("| " + " | ".join(proj['phases']) + " |")
        lines.append("")
        lines.append("---")
        lines.append("")

    # ── 業務実績サマリー（固定） ──────────────────────────────────
    lines.append("## ■ 業務実績サマリー（フリーランス）")
    lines.append("")
    lines.append("| 項目 | 内容 |")
    lines.append("| :--- | :--- |")
    lines.append("| **活動期間** | 2019年12月 〜 現在（約6年） |")
    lines.append("| **総受注数** | 174件（総プロジェクト数 183件） |")
    lines.append("| **顧客満足度** | 97%（高評価 178件 / 総評価 183件） |")
    lines.append("| **主な成果** | リサーチ業務の作業時間を **10時間 → 1時間（90%削減）** に短縮 |")
    lines.append("| **強み** | 要件定義〜納品後サポートまでの一貫対応、AIツールを活用した高速開発 |")
    lines.append("")

    return "\n".join(lines)


def main():
    print(f"1. {INPUT_XLSX} を読み込み中...")
    wb = openpyxl.load_workbook(INPUT_XLSX, data_only=True)
    ws = wb.active
    print(f"   シート: {ws.title} / 行: {ws.max_row} / 列: {ws.max_column}")

    profile = read_profile(ws)
    print(f"   氏名: {profile.get('氏名','(不明)')}")

    rows = read_all_rows(ws)
    projects = find_project_rows(rows)
    print(f"   プロジェクト数: {len(projects)} 件")
    for p in projects:
        print(f"   - {p['no']:>2}. {p['title'][:50]}")

    print(f"\n2. {OUTPUT_MD} を生成中...")
    md = build_md(profile, projects)
    with open(OUTPUT_MD, "w", encoding="utf-8") as f:
        f.write(md)
    print(f"   ✓ {OUTPUT_MD}")
    print("\n✅ 完了")


if __name__ == "__main__":
    main()
